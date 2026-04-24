import json
from openpyxl import load_workbook

INPUT_LINES = r'..\mappers\lines_master.xlsx'
OUTPUT_LINESTOSODOS = r'..\mappers\MASTERGO\MAPPED\lines_to_sodos_lines.json'


# ----------------------------
# Helpers
# ----------------------------

def safe_str(v):
    if v is None:
        return ""
    v = str(v).strip()
    if v.lower() in ("nan", "none", ""):
        return ""
    return v.lower()


def build_mapping(ws):
    headers = {}
    for idx, cell in enumerate(ws[1]):
        headers[safe_str(cell.value).lower()] = idx + 1

    def col(name):
        return headers.get(name.lower())

    line_col = col("line")
    map1_col = col("map1")
    map2_col = col("map2")
    direct_col = col("not_direct_match")

    if not line_col:
        raise ValueError("Missing required column: line")

    result = {}

    for row in range(2, ws.max_row + 1):

        line = safe_str(ws.cell(row=row, column=line_col).value)
        if not line:
            continue

        map1 = safe_str(ws.cell(row=row, column=map1_col).value)
        map2 = safe_str(ws.cell(row=row, column=map2_col).value)
        direct = safe_str(ws.cell(row=row, column=direct_col).value)

        direct = 1 if direct == "1" else 0

        if line not in result:
            result[line] = {
                "maps": [],
                "direct": 0
            }

        # add maps
        for m in (map1, map2):
            if m and m not in result[line]["maps"]:
                result[line]["maps"].append(m)

        # handle direct matches
        if direct:
            result[line]["direct"] = direct

    return result


# ----------------------------
# Main
# ----------------------------

def process():
    wb = load_workbook(INPUT_LINES)
    ws = wb.active

    data = build_mapping(ws)

    with open(OUTPUT_LINESTOSODOS, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"Saved mapping to: {OUTPUT_LINESTOSODOS}")


if __name__ == "__main__":
    process()